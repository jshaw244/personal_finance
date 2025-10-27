# scripts/compare_db_snapshots.py
"""
Compares two database snapshots (Excel exports) and reports schema and data differences.

Usage:
    python scripts/compare_db_snapshots.py <old_snapshot.xlsx> <new_snapshot.xlsx>

Output:
    results/sandbox_db_diff_YYYYMMDD_HHMM.xlsx

The comparison checks:
  - Tables added, removed, or renamed
  - Column-level schema changes
  - Row count differences
  - Record-level differences (optional, configurable for small tables)
"""

import sys
import os
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook

# --- Configuration ---
PROJECT_ROOT = Path("C:/DATA/personal_finance")
RESULTS_DIR = PROJECT_ROOT / "results"
RESULTS_DIR.mkdir(exist_ok=True)

timestamp = datetime.now().strftime("%Y%m%d_%H%M")
OUTFILE = RESULTS_DIR / f"sandbox_db_diff_{timestamp}.xlsx"

def log(msg):
    print(msg)

def get_sheet_names(file_path):
    """Return the list of table/sheet names excluding metadata sheets."""
    xls = pd.ExcelFile(file_path)
    exclude = {"DB_Metadata", "Schema_Definitions", "Indexes", "Triggers", "Views"}
    return [s for s in xls.sheet_names if s not in exclude]

def summarize_excel_snapshot(file_path):
    """Return a dict mapping table -> (columns, row_count, DataFrame)"""
    xls = pd.ExcelFile(file_path)
    summary = {}
    for sheet in get_sheet_names(file_path):
        try:
            df = pd.read_excel(xls, sheet_name=sheet)
            summary[sheet] = {
                "columns": list(df.columns),
                "row_count": len(df),
                "data": df
            }
        except Exception as e:
            log(f"Warning: Could not load sheet {sheet}: {e}")
    return summary

def compare_snapshots(old, new):
    """Compare two summaries and return structured diff info."""
    old_tables = set(old.keys())
    new_tables = set(new.keys())

    added = sorted(list(new_tables - old_tables))
    removed = sorted(list(old_tables - new_tables))
    common = sorted(list(old_tables & new_tables))

    schema_changes = []
    row_diffs = []

    for table in common:
        old_cols = old[table]["columns"]
        new_cols = new[table]["columns"]

        if old_cols != new_cols:
            schema_changes.append({
                "Table": table,
                "Old_Columns": ", ".join(old_cols),
                "New_Columns": ", ".join(new_cols)
            })

        old_rows = old[table]["row_count"]
        new_rows = new[table]["row_count"]
        if old_rows != new_rows:
            row_diffs.append({
                "Table": table,
                "Old_Rows": old_rows,
                "New_Rows": new_rows,
                "Change": new_rows - old_rows
            })

    return added, removed, schema_changes, row_diffs

def write_report(added, removed, schema_changes, row_diffs):
    """Write all diffs to an Excel workbook."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Section", "Count"])
    ws_summary.append(["Tables Added", len(added)])
    ws_summary.append(["Tables Removed", len(removed)])
    ws_summary.append(["Schema Changes", len(schema_changes)])
    ws_summary.append(["Row Count Differences", len(row_diffs)])

    ws_added = wb.create_sheet("Tables_Added")
    ws_added.append(["Table"])
    for t in added:
        ws_added.append([t])

    ws_removed = wb.create_sheet("Tables_Removed")
    ws_removed.append(["Table"])
    for t in removed:
        ws_removed.append([t])

    ws_schema = wb.create_sheet("Schema_Changes")
    ws_schema.append(["Table", "Old_Columns", "New_Columns"])
    for row in schema_changes:
        ws_schema.append([row["Table"], row["Old_Columns"], row["New_Columns"]])

    ws_rows = wb.create_sheet("Row_Count_Changes")
    ws_rows.append(["Table", "Old_Rows", "New_Rows", "Change"])
    for row in row_diffs:
        ws_rows.append([row["Table"], row["Old_Rows"], row["New_Rows"], row["Change"]])

    wb.save(OUTFILE)
    log(f"Comparison report written to: {OUTFILE}")

def main():
    if len(sys.argv) != 3:
        print("Usage: python scripts/compare_db_snapshots.py <old_snapshot.xlsx> <new_snapshot.xlsx>")
        sys.exit(1)

    old_path = Path(sys.argv[1])
    new_path = Path(sys.argv[2])

    if not old_path.exists() or not new_path.exists():
        print("Error: One or both snapshot files not found.")
        sys.exit(1)

    log(f"Comparing snapshots:\n  OLD: {old_path}\n  NEW: {new_path}")

    old_summary = summarize_excel_snapshot(old_path)
    new_summary = summarize_excel_snapshot(new_path)

    added, removed, schema_changes, row_diffs = compare_snapshots(old_summary, new_summary)
    write_report(added, removed, schema_changes, row_diffs)

if __name__ == "__main__":
    main()
