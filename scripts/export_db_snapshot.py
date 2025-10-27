# scripts/export_db_snapshot.py
"""
Exports the full SQLite sandbox database (data/plaid.db) into a timestamped Excel workbook.

Includes:
  - One sheet per table (full contents)
  - Schema definitions, indexes, triggers, and views
  - Metadata overview (row counts, sample data)
  - Auto-sanitization of illegal Excel characters
  - Logs any sanitization or export errors
  - Automatically zips and archives the final snapshot with SHA256 checksum

Output:
  results/sandbox_db_snapshot_YYYYMMDD_HHMM.xlsx
  archive/sandbox_db_snapshot_YYYYMMDD_HHMM.zip
"""

import os
import sqlite3
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from zipfile import ZipFile
import hashlib
import re

# --- Configuration ---
PROJECT_ROOT = Path("C:/DATA/personal_finance")
DB_PATH = PROJECT_ROOT / "data/plaid.db"
RESULTS_DIR = PROJECT_ROOT / "results"
ARCHIVE_DIR = PROJECT_ROOT / "archive"
LOGS_DIR = PROJECT_ROOT / "logs"

for d in (RESULTS_DIR, ARCHIVE_DIR, LOGS_DIR):
    d.mkdir(exist_ok=True)

timestamp = datetime.now().strftime("%Y%m%d_%H%M")
OUTFILE = RESULTS_DIR / f"sandbox_db_snapshot_{timestamp}.xlsx"
LOGFILE = LOGS_DIR / "db_snapshot_sanitizer.log"

def log(msg: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOGFILE, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {msg}\n")
    print(msg)

log(f"Starting SQLite export from {DB_PATH}")
log(f"Output file: {OUTFILE}")

# --- Utility: Safe sanitizer for Excel ---
def sanitize(value):
    """Remove illegal Excel characters while keeping printable content."""
    if isinstance(value, str):
        clean = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)
        if clean != value:
            log(f"Sanitized illegal characters in string: {value[:80]!r}")
        return clean
    return value

# --- Connect ---
conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

# --- Fetch tables & views ---
tables = cursor.execute(
    "SELECT name, type FROM sqlite_master WHERE type IN ('table','view') ORDER BY type, name"
).fetchall()

# --- Create workbook ---
wb = Workbook()
ws_meta = wb.active
ws_meta.title = "DB_Metadata"
ws_meta.append(["Name", "Type", "RowCount", "Columns", "Sample Data"])

# --- Export content ---
for name, obj_type in tables:
    try:
        df = pd.read_sql_query(f"SELECT * FROM {name}", conn)
        ws = wb.create_sheet(title=name[:31])
        ws.append(list(df.columns))

        for _, row in df.iterrows():
            ws.append([sanitize(v) for v in row])

        sample_val = str(df.iloc[0].to_dict())[:80] + "..." if len(df) > 0 else ""
        ws_meta.append([name, obj_type, len(df), len(df.columns), sample_val])

    except Exception as e:
        clean_msg = sanitize(str(e))
        ws = wb.create_sheet(title=f"{name[:28]}_ERR")
        ws.append(["Error", clean_msg])
        ws_meta.append([name, obj_type, "ERROR", "", clean_msg])
        log(f"Failed to export {name}: {clean_msg}")

# --- Schema Definitions ---
schema_df = pd.read_sql_query(
    "SELECT name, type, sql FROM sqlite_master WHERE sql IS NOT NULL ORDER BY type, name",
    conn,
)
ws_schema = wb.create_sheet(title="Schema_Definitions")
ws_schema.append(list(schema_df.columns))
for _, row in schema_df.iterrows():
    ws_schema.append([sanitize(v) for v in row])

# --- Indexes ---
indexes = cursor.execute(
    "SELECT name, tbl_name, sql FROM sqlite_master WHERE type='index' ORDER BY tbl_name"
).fetchall()
ws_indexes = wb.create_sheet(title="Indexes")
ws_indexes.append(["Index Name", "Table", "SQL"])
for name, tbl, sql in indexes:
    ws_indexes.append([sanitize(name), sanitize(tbl), sanitize(sql)])

# --- Triggers ---
triggers = cursor.execute(
    "SELECT name, tbl_name, sql FROM sqlite_master WHERE type='trigger' ORDER BY tbl_name"
).fetchall()
ws_triggers = wb.create_sheet(title="Triggers")
ws_triggers.append(["Trigger Name", "Table", "SQL"])
for name, tbl, sql in triggers:
    ws_triggers.append([sanitize(name), sanitize(tbl), sanitize(sql)])

# --- Views ---
views = cursor.execute(
    "SELECT name, sql FROM sqlite_master WHERE type='view' ORDER BY name"
).fetchall()
ws_views = wb.create_sheet(title="Views")
ws_views.append(["View Name", "SQL"])
for name, sql in views:
    ws_views.append([sanitize(name), sanitize(sql)])

conn.close()

# --- Save Excel file ---
wb.save(OUTFILE)
log(f"Excel snapshot created: {OUTFILE}")

# --- Compute SHA256 checksum ---
hash_sha256 = hashlib.sha256()
with open(OUTFILE, "rb") as f:
    for chunk in iter(lambda: f.read(4096), b""):
        hash_sha256.update(chunk)
checksum = hash_sha256.hexdigest()
log(f"SHA256 checksum: {checksum}")

# --- Create ZIP archive ---
zip_name = ARCHIVE_DIR / f"sandbox_db_snapshot_{timestamp}.zip"
with ZipFile(zip_name, "w") as zipf:
    zipf.write(OUTFILE, arcname=OUTFILE.name)
log(f"Archived snapshot: {zip_name}")

# --- Save checksum file alongside ZIP ---
with open(str(zip_name) + ".sha256", "w", encoding="utf-8") as f:
    f.write(f"{checksum}  {OUTFILE.name}\n")

log("Export, sanitization, and archival completed successfully.")
print(f"\nSnapshot complete -> {OUTFILE}\nArchived -> {zip_name}")
